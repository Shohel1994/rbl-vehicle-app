"""
Smart Vehicle Management System — Supabase (PostgreSQL) Edition
==================================================================
Enterprise-grade backend: Supabase Postgres via the official `supabase-py`
SDK, replacing SQLite. Adds self-service user registration with admin
approval, in addition to requisition approval. No email — everything
happens live inside the app. Exports: Excel (.xlsx) and PDF (.pdf).

Author: Senior Python Developer (generated for Shohel Rana)
"""

import io
import os
import base64
import hashlib
import random
import time
from datetime import datetime, date, timedelta
from secrets import token_urlsafe

import pandas as pd
import plotly.express as px
import streamlit as st
from supabase import create_client, Client
from fpdf import FPDF
from fpdf.enums import XPos, YPos
from fpdf.fonts import FontFace
from streamlit_autorefresh import st_autorefresh
import extra_streamlit_components as stx
import requests
import streamlit as st


# =========================================================
# TELEGRAM NOTIFICATION SYSTEM & HELPER FUNCTIONS
# =========================================================
def send_telegram_alert(message: str):
    """Sends real-time notifications to the Telegram group with secret fallback."""
    try:
        # st.secrets থেকে রিড করার সুরক্ষিত পদ্ধতি + ব্যাকআপ মান
        bot_token = st.secrets.get("TELEGRAM_BOT_TOKEN", "8868510704:AAGkOS_s70f7ARKpvLbP3OvDDNEzDcutqIY")
        chat_id = st.secrets.get("TELEGRAM_CHAT_ID", "-1004360578852")

        if not bot_token or not chat_id:
            return

        url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
        payload = {
            "chat_id": str(chat_id),
            "text": message,
            "parse_mode": "Markdown"
        }
        requests.post(url, json=payload, timeout=5)
    except Exception as e:
        print(f"Telegram Notification Error: {e}")


def insert_requisition(data: dict):
    sb = get_supabase_client()
    sb.table(REQUISITIONS_TABLE).insert(data).execute()

    # English Telegram Alert for New Requisition
    applicant = data.get("applicant_name", "N/A")
    dept = data.get("department", "N/A")
    dest = data.get("destination", "N/A")
    req_id = data.get("request_id", "N/A")
    date_of_travel = data.get("date_of_travel", "N/A")
    time_of_travel = data.get("time_of_travel", "N/A")
    vehicle_type = data.get("vehicle_type", "N/A")
    passenger_count = data.get("passenger_count", "N/A")

    msg = (
        f"🚨 **New Vehicle Requisition Submitted!**\n\n"
        f"🆔 **ID:** #{req_id}\n"
        f"👤 **Applicant:** {applicant}\n"
        f"🏢 **Department:** {dept}\n"
        f"📍 **Destination:** {dest}\n"
        f"📅 **Date/Time:** {date_of_travel} at {time_of_travel}\n"
        f"🚐 **Vehicle Type:** {vehicle_type}\n"
        f"👥 **Passengers:** {passenger_count}"
    )
    send_telegram_alert(msg)


def update_requisition(request_id: str, updates: dict):
    sb = get_supabase_client()

    # Look up applicant/destination so the alert is informative even though
    # `updates` itself usually only carries status-related fields.
    applicant = ""
    dest = ""
    try:
        existing = sb.table(REQUISITIONS_TABLE).select("applicant_name, destination").eq(
            "request_id", request_id
        ).limit(1).execute()
        if existing.data:
            applicant = existing.data[0].get("applicant_name", "")
            dest = existing.data[0].get("destination", "")
    except Exception:
        pass  # Alert enrichment is best-effort; the update itself must still proceed.

    sb.table(REQUISITIONS_TABLE).update(updates).eq("request_id", request_id).execute()

    # English Telegram Alert for Status Update
    status = updates.get("status", "Updated")
    driver = updates.get("driver_name", "")
    vehicle = updates.get("vehicle_number", "")

    msg = (
        f"📢 **Requisition Status Updated!**\n\n"
        f"🆔 **Requisition ID:** #{request_id}\n"
    )
    if applicant:
        msg += f"👤 **Applicant:** {applicant}\n"
    if dest:
        msg += f"📍 **Destination:** {dest}\n"
    msg += f"📌 **New Status:** {status}"
    if driver:
        msg += f"\n👨‍✈️ **Driver:** {driver}"
    if vehicle:
        msg += f"\n🚗 **Vehicle:** {vehicle}"

    send_telegram_alert(msg)


# =========================================================
# 1. COMPANY INFO & PAGE CONFIG
# =========================================================
COMPANY_NAME = "Renaissaince Barind Ltd."
COMPANY_ADDRESS = "Ishwardi EPZ, Pakshi, Pabna"

st.set_page_config(
    page_title="RBL VMS",
    page_icon="🚗",
    layout="wide",
    initial_sidebar_state="expanded",
)

USERS_TABLE = "users"
REQUISITIONS_TABLE = "requisitions"
DRIVERS_TABLE = "drivers"
VEHICLES_TABLE = "vehicles"
SESSIONS_TABLE = "sessions"

SESSION_COOKIE_NAME = "rbl_vms_session"
SESSION_LIFETIME_DAYS = 30

VEHICLE_TYPES = ["Private Car", "HIACE", "Pick-up Van", "Covered Van", "Truck", "Shipment Vehicle", "Other"]
DEPARTMENTS = [
    "Accounts", "Warehouse", "Factory Merchandising", "Commercial", "Floor Operation",
    "TSD", "QMS", "Production Planning & Control", "Cutting", "R & D", "Admin",
    "Technical", "Finishing", "Quality", "IE", "HR & Compliance", "Other",
]

# Account status (users table) — approval workflow for logins
USER_STATUS_OPTIONS = ["Pending", "Approved", "Rejected"]
ROLE_OPTIONS = ["user", "security_officer", "admin"]
ROLE_DISPLAY = {"user": "Employee", "security_officer": "Security Officer", "admin": "Admin"}

# Requisition status (requisitions table) — full trip lifecycle
REQ_STATUS_OPTIONS = ["Pending", "Approved", "Rejected", "On Trip", "Completed"]
STATUS_BADGE = {
    "Pending": "🟡 Pending", "Approved": "🟢 Approved", "Rejected": "🔴 Rejected",
    "On Trip": "🔵 On Trip", "Completed": "✅ Completed",
}


# =========================================================
# 2. HELPER FUNCTIONS & HEADER (No Logo Version)
# =========================================================
@st.cache_data(show_spinner=False)
def get_logo_base64():
    """Logo system completely disabled."""
    return None


def company_header(subtitle: str = ""):
    """Reusable company name + address banner without logo."""
    st.title(f"🚗 {COMPANY_NAME}")
    st.caption(f"📍 {COMPANY_ADDRESS}")
    if subtitle:
        st.subheader(subtitle)
    st.divider()


def badge_class(status: str) -> str:
    return {
        "Pending": "badge-pending", "Approved": "badge-approved", "Rejected": "badge-rejected",
        "On Trip": "badge-ontrip", "Completed": "badge-completed",
    }.get(status, "badge-pending")


def is_blank(val) -> bool:
    if val is None:
        return True
    if isinstance(val, float) and pd.isna(val):
        return True
    if isinstance(val, str) and not val.strip():
        return True
    return False


def fmt(val, default: str = "—") -> str:
    return default if is_blank(val) else str(val)


def hash_password(raw: str) -> str:
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


# =========================================================
# 3. SUPABASE CONNECTION
# =========================================================
@st.cache_resource(show_spinner="Connecting to Supabase...")
def get_supabase_client() -> Client:
    url = st.secrets["SUPABASE_URL"]
    key = st.secrets["SUPABASE_KEY"]
    return create_client(url, key)


def check_tables_ready() -> tuple[bool, str]:
    """Verify the users/requisitions tables exist and are reachable."""
    sb = get_supabase_client()
    try:
        sb.table(USERS_TABLE).select("id").limit(1).execute()
        sb.table(REQUISITIONS_TABLE).select("id").limit(1).execute()
        return True, ""
    except Exception as e:
        return False, str(e)


# ---------------------- USERS TABLE HELPERS ----------------------
def get_user_by_username(username: str):
    sb = get_supabase_client()
    res = sb.table(USERS_TABLE).select("*").eq("username", username).limit(1).execute()
    return res.data[0] if res.data else None


def register_user(data: dict):
    sb = get_supabase_client()
    sb.table(USERS_TABLE).insert(data).execute()


def update_user(username: str, updates: dict):
    sb = get_supabase_client()
    sb.table(USERS_TABLE).update(updates).eq("username", username).execute()


def delete_user(username: str):
    """Permanently remove a user account (e.g. after they leave the company)."""
    sb = get_supabase_client()
    sb.table(USERS_TABLE).delete().eq("username", username).execute()


def fetch_all_users() -> pd.DataFrame:
    sb = get_supabase_client()
    res = sb.table(USERS_TABLE).select("*").order("created_at", desc=True).execute()
    return pd.DataFrame(res.data) if res.data else pd.DataFrame(
        columns=["id", "username", "full_name", "designation", "employee_id", "department",
                 "mobile", "role", "status", "created_at"]
    )


# ------------------- REQUISITIONS TABLE HELPERS -------------------
def generate_request_id() -> str:
    return f"REQ-{datetime.now().strftime('%Y%m%d%H%M%S')}-{random.randint(100, 999)}"


# NOTE: insert_requisition() and update_requisition() are defined once, above,
# in the "TELEGRAM NOTIFICATION SYSTEM & HELPER FUNCTIONS" section — they
# perform the Supabase write AND fire the Telegram alert. They are
# intentionally not redefined here; a second, alert-less definition at this
# point in the file would silently shadow (override) the ones above and the
# Telegram notifications would never fire, since Python keeps whichever `def`
# runs last for a given name.


def fetch_all_requisitions() -> pd.DataFrame:
    sb = get_supabase_client()
    res = sb.table(REQUISITIONS_TABLE).select("*").order("created_at", desc=True).execute()
    return pd.DataFrame(res.data) if res.data else pd.DataFrame(columns=[
        "id", "request_id", "created_at", "username", "applicant_name", "department", "mobile_number",
        "date_of_travel", "time_of_travel", "destination", "passenger_count", "vehicle_type", "purpose",
        "special_request", "status", "driver_name", "driver_contact", "vehicle_number", "approved_by",
        "action_timestamp", "approved_time", "admin_note", "start_km", "end_km", "total_km",
        "actual_exit_time", "actual_return_time",
    ])


def fetch_requisitions_by_user(username: str) -> pd.DataFrame:
    """Server-side filtered fetch — strict data isolation: only this user's rows are ever requested."""
    sb = get_supabase_client()
    res = (
        sb.table(REQUISITIONS_TABLE)
        .select("*")
        .eq("username", username)
        .order("created_at", desc=True)
        .execute()
    )
    return pd.DataFrame(res.data) if res.data else pd.DataFrame(columns=[
        "id", "request_id", "created_at", "username", "applicant_name", "department", "mobile_number",
        "date_of_travel", "time_of_travel", "destination", "passenger_count", "vehicle_type", "purpose",
        "special_request", "status", "driver_name", "driver_contact", "vehicle_number", "approved_by",
        "action_timestamp", "approved_time", "admin_note", "start_km", "end_km", "total_km",
        "actual_exit_time", "actual_return_time",
    ])


def fetch_requisitions_by_status(status: str) -> pd.DataFrame:
    """Server-side filtered fetch used by the Security Officer panel — only pulls
    requisitions in the given trip-status (e.g. 'Approved' or 'On Trip'), so
    Pending/Rejected requests are never even requested, let alone shown."""
    sb = get_supabase_client()
    res = (
        sb.table(REQUISITIONS_TABLE)
        .select("*")
        .eq("status", status)
        .order("created_at", desc=True)
        .execute()
    )
    return pd.DataFrame(res.data) if res.data else pd.DataFrame(columns=[
        "id", "request_id", "created_at", "username", "applicant_name", "department", "mobile_number",
        "date_of_travel", "time_of_travel", "destination", "passenger_count", "vehicle_type", "purpose",
        "special_request", "status", "driver_name", "driver_contact", "vehicle_number", "approved_by",
        "action_timestamp", "approved_time", "admin_note", "start_km", "end_km", "total_km",
        "actual_exit_time", "actual_return_time",
    ])


# ------------------- DRIVERS & VEHICLES TABLE HELPERS -------------------
# These back the dynamic dropdowns in the requisition-approval form so admins
# maintain one source of truth instead of retyping names/numbers each time.
def fetch_all_drivers() -> pd.DataFrame:
    sb = get_supabase_client()
    res = sb.table(DRIVERS_TABLE).select("*").order("driver_name").execute()
    return pd.DataFrame(res.data) if res.data else pd.DataFrame(columns=["id", "driver_name", "driver_contact", "created_at"])


def add_driver(driver_name: str, driver_contact: str):
    sb = get_supabase_client()
    sb.table(DRIVERS_TABLE).insert({"driver_name": driver_name, "driver_contact": driver_contact}).execute()


def delete_driver(driver_id):
    sb = get_supabase_client()
    sb.table(DRIVERS_TABLE).delete().eq("id", driver_id).execute()


def fetch_all_vehicles() -> pd.DataFrame:
    sb = get_supabase_client()
    res = sb.table(VEHICLES_TABLE).select("*").order("vehicle_number").execute()
    return pd.DataFrame(res.data) if res.data else pd.DataFrame(columns=["id", "vehicle_number", "created_at"])


def add_vehicle(vehicle_number: str):
    sb = get_supabase_client()
    sb.table(VEHICLES_TABLE).insert({"vehicle_number": vehicle_number}).execute()


def delete_vehicle(vehicle_id):
    sb = get_supabase_client()
    sb.table(VEHICLES_TABLE).delete().eq("id", vehicle_id).execute()


# ------------------- SESSION (REMEMBER ME) HELPERS -------------------
# A "remember me" cookie stores only an opaque, unguessable token — never the
# username or password directly — so a leaked/inspected cookie can't be used
# to reconstruct credentials. The token maps to a username via this table and
# expires automatically, and is revoked (deleted) on explicit logout.
def create_session(username: str) -> str:
    sb = get_supabase_client()
    token = token_urlsafe(32)
    expires_at = (datetime.utcnow() + timedelta(days=SESSION_LIFETIME_DAYS)).strftime("%Y-%m-%d %H:%M:%S")
    sb.table(SESSIONS_TABLE).insert({
        "token": token, "username": username, "expires_at": expires_at,
    }).execute()
    return token


def get_session_username(token: str):
    """Return the username for a still-valid session token, or None."""
    if not token:
        return None
    sb = get_supabase_client()
    res = sb.table(SESSIONS_TABLE).select("*").eq("token", token).limit(1).execute()
    if not res.data:
        return None
    row = res.data[0]
    try:
        # PostgREST reads timestamptz columns back in ISO 8601 with a 'T'
        # separator and often a timezone offset (e.g. "...T14:22:30.123+00:00"),
        # which differs from the space-separated string we wrote on insert —
        # normalize both shapes before parsing so expiry checks don't silently
        # fail (and reject) every real session.
        raw = row["expires_at"].replace("T", " ")
        expires_at = datetime.strptime(raw[:19], "%Y-%m-%d %H:%M:%S")
    except (ValueError, TypeError, KeyError, AttributeError):
        return None
    if expires_at < datetime.utcnow():
        return None
    return row["username"]


def delete_session(token: str):
    if not token:
        return
    sb = get_supabase_client()
    sb.table(SESSIONS_TABLE).delete().eq("token", token).execute()


def get_cookie_manager():
    """Returns a CookieManager. Deliberately NOT wrapped in st.cache_resource:
    that cache is shared globally across every visitor's session on the
    server, and caching a stateful per-browser cookie wrapper there would
    leak one user's session cookie into another user's script run. Streamlit's
    component protocol is already session-scoped on its own, so a fresh,
    cheap instantiation on every rerun is the correct (and documented)
    pattern for this library."""
    return stx.CookieManager(key="rbl_vms_cookie_manager")


# =========================================================
# 4. AUTH: SIGN IN + SELF REGISTRATION
# =========================================================
def get_bootstrap_admin():
    """Fallback super-admin defined in secrets, used only to bootstrap the very
    first login before any 'admin' role exists in the users table."""
    try:
        return st.secrets["admin"]["username"], st.secrets["admin"]["password"]
    except Exception:
        return None, None


def build_bootstrap_user_dict(username: str) -> dict:
    return {"username": username, "full_name": "Super Admin (Bootstrap)", "role": "admin",
            "department": "Management", "designation": "System Administrator", "mobile": ""}


def build_user_dict(record: dict) -> dict:
    """Shared by both password login and cookie-based session restore, so the
    session_state.auth_user shape never drifts between the two paths."""
    return {
        "username": record["username"],
        "full_name": record.get("full_name", record["username"]),
        "role": record.get("role", "user"),
        "department": record.get("department", ""),
        "designation": record.get("designation", ""),
        "mobile": record.get("mobile", ""),
    }


def attempt_login(username: str, password: str):
    boot_user, boot_pass = get_bootstrap_admin()
    if boot_user and username == boot_user and password == boot_pass:
        return build_bootstrap_user_dict(username)

    record = get_user_by_username(username)
    if not record:
        return None
    if record.get("status") != "Approved":
        st.error(f"⏳ Your account status is **{record.get('status', 'Pending')}**. Please wait for admin approval.")
        return "PENDING"
    if record.get("password") != hash_password(password):
        return None
    return build_user_dict(record)


def restore_user_from_username(username: str):
    """Used only for cookie-based 'remember me' restore — re-validates the
    account is still Approved (in case it was later revoked/rejected) before
    trusting the session token."""
    boot_user, _ = get_bootstrap_admin()
    if boot_user and username == boot_user:
        return build_bootstrap_user_dict(username)
    record = get_user_by_username(username)
    if record and record.get("status") == "Approved":
        return build_user_dict(record)
    return None


def login_view():
    company_header("Smart Vehicle Management System — Sign in or request a new account")
    ready, err = check_tables_ready()
    if not ready:
        st.error(
            "⚠️ Could not reach the `users` / `requisitions` tables in Supabase. "
            "Make sure you've run the setup SQL from `SETUP_GUIDE.md` in the Supabase SQL Editor, "
            "and that `SUPABASE_URL` / `SUPABASE_KEY` in secrets are correct."
        )
        with st.expander("Technical details"):
            st.code(err)
        st.stop()

    _, mid, _ = st.columns([1, 1.3, 1])
    with mid:
        tab_login, tab_register = st.tabs(["🔐 Sign In", "📝 Request New User ID"])

        with tab_login:
            with st.form("login_form"):
                username = st.text_input("Username")
                password = st.text_input("Password", type="password")
                remember_me = st.checkbox("Remember me on this device", value=True)
                submitted = st.form_submit_button("Login", type="primary", use_container_width=True)

            if submitted:
                result = attempt_login(username.strip(), password)
                if result == "PENDING":
                    pass  # error already shown
                elif result is None:
                    st.error("❌ Invalid username or password.")
                else:
                    session_token = None
                    if remember_me:
                        try:
                            session_token = create_session(result["username"])
                            cookie_manager.set(
                                SESSION_COOKIE_NAME, session_token, key="set_login_cookie",
                                expires_at=datetime.now() + timedelta(days=SESSION_LIFETIME_DAYS),
                            )
                            # IMPORTANT: cookie_manager.set() only *dispatches* a message
                            # to the browser's cookie-manager iframe component asking it
                            # to write document.cookie — it does not write synchronously.
                            # If we st.rerun() immediately, Streamlit tears the page down
                            # before the browser JS has a chance to actually persist the
                            # cookie, so "Remember me" silently never works. A short pause
                            # here gives the component time to finish the write before the
                            # rerun happens.
                            with st.spinner("Setting up your session..."):
                                time.sleep(1.0)
                        except Exception:
                            session_token = None  # Remember-me is best-effort; login still succeeds without it.
                    result["session_token"] = session_token
                    st.session_state.auth_user = result
                    st.rerun()

        with tab_register:
            st.caption("New employees or Security Officers can request an account here. "
                       "An admin must approve your account before you can log in.")

            reg_role = st.radio(
                "Register as", ["Employee", "Security Officer"],
                horizontal=True, key="reg_role_choice",
            )
            st.markdown("---")

            if reg_role == "Employee":
                st.caption("Your default password will be your **Employee ID**.")
                with st.form("register_employee_form", clear_on_submit=True):
                    full_name = st.text_input("Full Name *")
                    username_r = st.text_input("Choose a Username *")
                    designation = st.text_input("Designation *")
                    employee_id = st.text_input("Employee ID *", help="This will also be your default password.")
                    department = st.selectbox("Department *", DEPARTMENTS)
                    mobile = st.text_input("Mobile Number *")
                    reg_submitted = st.form_submit_button("Submit Request", type="primary", use_container_width=True)

                if reg_submitted:
                    errors = []
                    if not full_name.strip():
                        errors.append("Full Name is required.")
                    if not username_r.strip():
                        errors.append("Username is required.")
                    if not employee_id.strip():
                        errors.append("Employee ID is required.")
                    if not mobile.strip():
                        errors.append("Mobile Number is required.")
                    if not errors and get_user_by_username(username_r.strip()):
                        errors.append("This username is already taken. Please choose another.")

                    if errors:
                        for e in errors:
                            st.error(e)
                    else:
                        register_user({
                            "username": username_r.strip(),
                            "password": hash_password(employee_id.strip()),
                            "full_name": full_name.strip(),
                            "designation": designation.strip(),
                            "employee_id": employee_id.strip(),
                            "department": department,
                            "mobile": mobile.strip(),
                            "role": "user",
                            "status": "Pending",
                        })
                        st.success(
                            "✅ Your account request has been submitted! Your default password is your "
                            f"**Employee ID ({employee_id.strip()})**. Please wait for admin approval before signing in."
                        )

            else:  # Security Officer registration — same simple flow as normal users
                st.caption("Choose your own username and password below, just like a regular account.")
                with st.form("register_security_form", clear_on_submit=True):
                    sec_full_name = st.text_input("Full Name *", key="sec_full_name")
                    sec_username = st.text_input("Username *", key="sec_username")
                    sec_password = st.text_input("Password *", type="password", key="sec_password")
                    sec_password_confirm = st.text_input("Confirm Password *", type="password", key="sec_password_confirm")
                    sec_submitted = st.form_submit_button("Submit Request", type="primary", use_container_width=True)

                if sec_submitted:
                    errors = []
                    if not sec_full_name.strip():
                        errors.append("Full Name is required.")
                    if not sec_username.strip():
                        errors.append("Username is required.")
                    if not sec_password:
                        errors.append("Password is required.")
                    elif len(sec_password) < 4:
                        errors.append("Password must be at least 4 characters.")
                    elif sec_password != sec_password_confirm:
                        errors.append("Passwords do not match.")
                    if not errors and get_user_by_username(sec_username.strip()):
                        errors.append("This username is already taken. Please choose another.")

                    if errors:
                        for e in errors:
                            st.error(e)
                    else:
                        register_user({
                            "username": sec_username.strip(),
                            "password": hash_password(sec_password),
                            "full_name": sec_full_name.strip(),
                            "designation": "Security Officer",
                            "employee_id": "",
                            "department": "Security",
                            "mobile": "",
                            "role": "security_officer",
                            "status": "Pending",
                        })
                        st.success(
                            "✅ Your Security Officer account request has been submitted! "
                            "Please wait for admin approval before signing in."
                        )


def logout_button():
    if st.sidebar.button("🚪 Logout", use_container_width=True):
        token = st.session_state.get("auth_user", {}).get("session_token")
        if token:
            try:
                delete_session(token)
            except Exception:
                pass  # DB cleanup is best-effort; logout must still proceed either way.
        try:
            cookie_manager.delete(SESSION_COOKIE_NAME, key="delete_logout_cookie")
        except KeyError:
            pass  # No cookie was ever set for this session — nothing to remove.
        del st.session_state["auth_user"]
        st.rerun()


# =========================================================
# 5. EXCEL + PDF REPORT GENERATORS
# =========================================================
def build_excel_report(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Requisitions")
    return buf.getvalue()


class ReportPDF(FPDF):
    def header(self):
        self.set_font("Helvetica", "B", 16)
        self.set_text_color(15, 98, 254)
        self.cell(0, 9, COMPANY_NAME, align="C", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

        self.set_font("Helvetica", "", 10)
        self.set_text_color(90, 90, 90)
        self.cell(0, 6, COMPANY_ADDRESS, align="C", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

        self.set_font("Helvetica", "B", 12)
        self.set_text_color(0, 0, 0)
        self.cell(0, 8, "Vehicle Requisition Report", align="C", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        self.ln(2)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(150, 150, 150)
        self.cell(0, 10, f"Page {self.page_no()}", align="C")


def build_pdf_report(df: pd.DataFrame, filters_summary: str) -> bytes:
    pdf = ReportPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Helvetica", size=10)
    pdf.cell(0, 6, f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
              new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.multi_cell(0, 6, filters_summary)
    pdf.ln(2)

    total = len(df)
    approved = int((df["status"] == "Approved").sum()) if not df.empty else 0
    rejected = int((df["status"] == "Rejected").sum()) if not df.empty else 0
    pending = int((df["status"] == "Pending").sum()) if not df.empty else 0
    on_trip = int((df["status"] == "On Trip").sum()) if not df.empty else 0
    completed = int((df["status"] == "Completed").sum()) if not df.empty else 0

    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, "Summary Statistics", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Helvetica", size=9)
    pdf.set_fill_color(230, 230, 230)
    for label, value in [("Total", total), ("Pending", pending), ("Approved", approved),
                          ("On Trip", on_trip), ("Completed", completed), ("Rejected", rejected)]:
        pdf.cell(38, 7, f"{label}: {value}", border=1, align="C", fill=True)
    pdf.ln(12)

    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, "Requisition Records", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(1)

    # Column widths are sized generously for "Req ID" (the longest field, e.g.
    # "REQ-20260809072230-852") and every cell wraps automatically via fpdf2's
    # table() API — this is what actually prevents text from bleeding into the
    # next column, instead of truncating with "…" as the previous version did.
    headers = ["Req ID", "Applicant", "Department", "Date", "Time", "Destination",
               "Vehicle", "Status", "Driver", "Vehicle No.", "Total KM"]
    cols = ["request_id", "applicant_name", "department", "date_of_travel", "time_of_travel",
            "destination", "vehicle_type", "status", "driver_name", "vehicle_number", "total_km"]
    col_widths = [42, 24, 22, 18, 13, 24, 16, 16, 22, 28, 16]

    pdf.set_font("Helvetica", size=7)
    heading_style = FontFace(emphasis="BOLD", color=(255, 255, 255), fill_color=(15, 98, 254))
    with pdf.table(col_widths=col_widths, text_align="LEFT", first_row_as_headings=True,
                   line_height=5, headings_style=heading_style, cell_fill_color=(245, 245, 245),
                   cell_fill_mode="ROWS") as table:
        header_row = table.row()
        for h in headers:
            header_row.cell(h)
        for _, r in df.iterrows():
            row = table.row()
            for col in cols:
                row.cell(fmt(r.get(col, ""), ""))

    return bytes(pdf.output())


# =========================================================
# 5B. DUTY TRACKER — EXCEL + PDF REPORT GENERATORS
# =========================================================
# These are ADDITIVE helpers for the new "Duty Tracker & Analytics" tab
# (Section 10B below). They are intentionally separate from
# build_excel_report() / build_pdf_report() / ReportPDF above so the
# existing "All Requisitions & Export" tab (Tab 5) keeps behaving exactly
# as before — nothing here is called from, or changes, that code path.

DUTY_TRACKER_DISPLAY_COLS = [
    "Vehicle No", "Driver Name", "Start Time", "End Time",
    "Total KM", "Duty Duration (Hrs)", "Route / Purpose",
]


def sanitize_pdf_text(value) -> str:
    """Make any string safe to hand to FPDF's core 'Helvetica' font.

    Core (non-embedded) PDF fonts like Helvetica only support the Latin-1
    character set. Common "smart" punctuation that Python/pandas/Streamlit
    happily display — em dashes (—), en dashes (–), curly quotes ('' ""),
    ellipses (…), bullets (•) — falls outside that set and makes fpdf2 raise
    FPDFUnicodeEncodingException the moment it's written to a cell. This
    function swaps the common offenders for plain-ASCII equivalents, then
    uses a Latin-1 encode/decode round-trip as a final safety net so any
    other unsupported character (e.g. stray emoji, non-Latin scripts such as
    Bangla typed into a destination/purpose field) degrades to '?' instead of
    crashing the whole export.

    Note: this keeps the export crash-proof but Latin-1-only. If admin notes,
    destinations, or driver names need full Bangla/Unicode rendering in the
    PDF itself, that requires embedding a Unicode TTF font (e.g. via
    pdf.add_font(...)) instead of the core Helvetica font — a larger change,
    out of scope for this fix.
    """
    if value is None:
        return ""
    text = str(value)
    replacements = {
        "\u2014": "-",   # — em dash
        "\u2013": "-",   # – en dash
        "\u2015": "-",   # ― horizontal bar
        "\u2018": "'", "\u2019": "'",   # ‘ ’ curly single quotes
        "\u201c": '"', "\u201d": '"',   # “ ” curly double quotes
        "\u2026": "...",  # … ellipsis
        "\u2022": "-",   # • bullet
        "\u2192": "->",  # → arrow
        "\u00a0": " ",   # non-breaking space
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    # Final safety net: anything still outside Latin-1 becomes '?' rather
    # than raising, so the PDF always generates successfully.
    return text.encode("latin-1", "replace").decode("latin-1")


def build_duty_tracker_excel(detail_df: pd.DataFrame, summary_metrics: dict) -> bytes:
    """Formatted .xlsx export for the Duty Tracker: a 'Summary' sheet with the
    KPI cards' values, plus a 'Duty Log' sheet with the full filtered detail
    table. Plug in your own detail_df / summary_metrics from the tab below —
    both are plain pandas / dict objects, nothing Supabase-specific here.
    """
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        summary_df = pd.DataFrame(
            [{"Metric": k, "Value": v} for k, v in summary_metrics.items()]
        )
        summary_df.to_excel(writer, index=False, sheet_name="Summary")
        detail_df.to_excel(writer, index=False, sheet_name="Duty Log")

        # Light auto-fit so columns aren't clipped in Excel — purely cosmetic,
        # safe to remove if you don't want the extra openpyxl dependency calls.
        from openpyxl.utils import get_column_letter
        for sheet_name, sheet_df in (("Summary", summary_df), ("Duty Log", detail_df)):
            ws = writer.sheets[sheet_name]
            for i, col in enumerate(sheet_df.columns, start=1):
                width = max(12, min(40, int(sheet_df[col].astype(str).map(len).max() if not sheet_df.empty else 12) + 2))
                ws.column_dimensions[get_column_letter(i)].width = width

    return buf.getvalue()


class DutyTrackerPDF(FPDF):
    """Separate FPDF subclass (rather than reusing ReportPDF) so this report's
    title/branding can evolve independently of the existing requisition report."""

    def header(self):
        self.set_font("Helvetica", "B", 16)
        self.set_text_color(15, 98, 254)
        self.cell(0, 9, COMPANY_NAME, align="C", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

        self.set_font("Helvetica", "", 10)
        self.set_text_color(90, 90, 90)
        self.cell(0, 6, COMPANY_ADDRESS, align="C", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

        self.set_font("Helvetica", "B", 12)
        self.set_text_color(0, 0, 0)
        self.cell(0, 8, "Vehicle & Driver Duty Tracker Report", align="C", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        self.ln(2)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(150, 150, 150)
        self.cell(0, 10, f"Page {self.page_no()}", align="C")


def build_duty_tracker_pdf(detail_df: pd.DataFrame, summary_metrics: dict, filters_summary: str) -> bytes:
    """PDF containing the KPI summary table followed by the detailed duty log.
    `detail_df` must already have the DUTY_TRACKER_DISPLAY_COLS columns (see
    the tab below for how it's built from the requisitions DataFrame).

    Every string written to the PDF is passed through sanitize_pdf_text()
    first — see that function's docstring for why this is necessary with
    FPDF's core Helvetica font.
    """
    pdf = DutyTrackerPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Helvetica", size=10)
    pdf.cell(0, 6, f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
              new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.multi_cell(0, 6, sanitize_pdf_text(filters_summary))
    pdf.ln(2)

    # ---- KPI Summary block ----
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, "Summary Metrics", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Helvetica", size=9)
    pdf.set_fill_color(230, 230, 230)
    for label, value in summary_metrics.items():
        safe_label = sanitize_pdf_text(label)
        safe_value = sanitize_pdf_text(value)
        pdf.cell(0, 7, f"{safe_label}: {safe_value}", border=1, fill=True, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(6)

    # ---- Detailed duty log table ----
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, "Detailed Duty Log", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(1)

    headers = DUTY_TRACKER_DISPLAY_COLS
    col_widths = [30, 28, 34, 34, 18, 26, 107]  # sums to ~277mm, fits A4 landscape

    pdf.set_font("Helvetica", size=7)
    heading_style = FontFace(emphasis="BOLD", color=(255, 255, 255), fill_color=(15, 98, 254))
    with pdf.table(col_widths=col_widths, text_align="LEFT", first_row_as_headings=True,
                   line_height=5, headings_style=heading_style, cell_fill_color=(245, 245, 245),
                   cell_fill_mode="ROWS") as table:
        header_row = table.row()
        for h in headers:
            header_row.cell(sanitize_pdf_text(h))
        for _, r in detail_df.iterrows():
            row = table.row()
            for col in headers:
                row.cell(sanitize_pdf_text(fmt(r.get(col, ""), "")))

    return bytes(pdf.output())


# =========================================================
# 6. SESSION STATE INIT (with "Remember Me" cookie restore)
# =========================================================
# Instantiated exactly once per script run — the underlying component uses a
# fixed key, and Streamlit errors on duplicate keys within a single run, so
# every other place in this file that needs cookies reuses this same object
# rather than calling get_cookie_manager() again.
cookie_manager = get_cookie_manager()

if "auth_user" not in st.session_state:
    # extra_streamlit_components's CookieManager runs inside a browser iframe
    # component. On the very first script run after a fresh page load, that
    # component hasn't finished round-tripping "here are the browser's
    # cookies" back to Python yet — cookie_manager.get_all() returns None
    # (NOT an empty dict) while it's still waiting. If we treat that None as
    # "no cookie exists" and immediately show the login form, the remembered
    # session is thrown away on every single page load, which is exactly the
    # "Remember me doesn't work" symptom. An empty dict {} — as opposed to
    # None — genuinely means "component is ready, browser has no cookie".
    all_cookies = cookie_manager.get_all()

    if all_cookies is None and not st.session_state.get("_cookie_bootstrap_done"):
        st.session_state["_cookie_bootstrap_done"] = True
        # Give the cookie component a brief moment to finish loading, then
        # force exactly one rerun so we re-check with real cookie data.
        st_autorefresh(interval=300, limit=1, key="cookie_bootstrap_refresh")
        st.stop()

    restored_user = None
    session_token = (all_cookies or {}).get(SESSION_COOKIE_NAME)
    if session_token:
        remembered_username = get_session_username(session_token)
        if remembered_username:
            restored_user = restore_user_from_username(remembered_username)
        if restored_user:
            restored_user["session_token"] = session_token
            st.session_state.auth_user = restored_user
        else:
            # Token is missing/expired/revoked, or the account is no longer
            # Approved — clear the stale cookie so we don't keep retrying it.
            try:
                cookie_manager.delete(SESSION_COOKIE_NAME, key="delete_stale_cookie")
            except KeyError:
                pass

    if "auth_user" not in st.session_state:
        login_view()
        st.stop()

user = st.session_state.auth_user

# =========================================================
# 7. SIDEBAR
# =========================================================
logo_uri = get_logo_base64()
logo_img = f'<img src="{logo_uri}" alt="logo">' if logo_uri else ""
st.sidebar.markdown(
    f'<div class="sidebar-brand">{logo_img}<span>{COMPANY_NAME}</span></div>',
    unsafe_allow_html=True,
)
st.sidebar.caption(f"📍 {COMPANY_ADDRESS}")
st.sidebar.markdown("---")
st.sidebar.markdown(f"**{user['full_name']}**")
st.sidebar.caption(f"Role: {ROLE_DISPLAY.get(user['role'], user['role'].capitalize())}")

auto_refresh_on = st.sidebar.checkbox("🔄 Auto-refresh every 10s", value=True,
                                       help="Automatically reloads live data across the app. "
                                            "Turn off temporarily if you're filling out a long form.")
if st.sidebar.button("🔄 Refresh Now", use_container_width=True):
    st.rerun()
st.sidebar.markdown("---")
logout_button()

if auto_refresh_on:
    st_autorefresh(interval=10_000, key="global_autorefresh")

# =========================================================
# 8. EMPLOYEE DASHBOARD
# =========================================================
if user["role"] == "user":
    company_header("👤 Employee Dashboard")
    tab1, tab2 = st.tabs(["📋 New Requisition", "📍 My Requests / Live Status"])

    with tab1:
        st.subheader("Submit a New Vehicle Requisition")
        with st.form("req_form", clear_on_submit=True):
            c1, c2 = st.columns(2)
            with c1:
                applicant_name = st.text_input("Applicant Name *", value=user["full_name"])
                department = st.selectbox("Department *", DEPARTMENTS,
                                           index=DEPARTMENTS.index(user["department"]) if user.get("department") in DEPARTMENTS else 0)
                mobile_number = st.text_input("Mobile Number *", value=user.get("mobile", ""),
                                               placeholder="01XXXXXXXXX")
                passenger_count = st.number_input("Passenger Count *", min_value=1, max_value=50, value=1)
            with c2:
                date_of_travel = st.date_input("Date of Travel *", min_value=date.today())
                time_of_travel = st.time_input("Time of Travel *")
                destination = st.text_input("Destination *")
                vehicle_type = st.selectbox("Vehicle Type Required *", VEHICLE_TYPES)

            purpose = st.text_area("Purpose of Travel *", height=90)
            special_request = st.text_area("Special Request (optional)", height=70)
            submitted = st.form_submit_button("🚀 Submit Requisition", type="primary", use_container_width=True)

        if submitted:
            errors = []
            if not applicant_name.strip():
                errors.append("Applicant Name is required.")
            if not mobile_number.strip():
                errors.append("Mobile Number is required.")
            if not destination.strip():
                errors.append("Destination is required.")
            if not purpose.strip():
                errors.append("Purpose of Travel is required.")

            if errors:
                for e in errors:
                    st.error(e)
            else:
                request_id = generate_request_id()
                data = {
                    "request_id": request_id,
                    "username": user["username"],
                    "applicant_name": applicant_name.strip(),
                    "department": department,
                    "mobile_number": mobile_number.strip(),
                    "date_of_travel": str(date_of_travel),
                    "time_of_travel": time_of_travel.strftime("%H:%M"),
                    "destination": destination.strip(),
                    "passenger_count": int(passenger_count),
                    "vehicle_type": vehicle_type,
                    "purpose": purpose.strip(),
                    "special_request": special_request.strip(),
                    "status": "Pending",
                    "driver_name": "",
                    "driver_contact": "",
                    "vehicle_number": "",
                    "approved_by": "",
                }
                with st.spinner("Saving to Supabase..."):
                    try:
                        insert_requisition(data)
                        st.success(f"✅ Requisition submitted! Your Request ID is **{request_id}**")
                        st.balloons()
                    except Exception as e:
                        st.error(f"❌ Failed to save requisition: {e}")

    with tab2:
        st.subheader("My Requests — Live Status")
        with st.spinner("Loading your requests..."):
            my_df = fetch_requisitions_by_user(user["username"])
        if my_df.empty:
            st.info("You haven't submitted any requisitions yet.")
        else:
            for _, r in my_df.iterrows():
                with st.container():
                    st.markdown(f"""
                    <div class="req-card">
                        <b>{r['request_id']}</b> &nbsp;|&nbsp; {r['destination']} &nbsp;|&nbsp;
                        {r['date_of_travel']} at {r['time_of_travel']} &nbsp;&nbsp;
                        <span class="{badge_class(r['status'])}">{STATUS_BADGE.get(r['status'], r['status'])}</span>
                    </div>
                    """, unsafe_allow_html=True)
                    with st.expander("View details"):
                        st.write(f"**Department:** {r['department']}  |  **Vehicle Type:** {r['vehicle_type']}  |  **Passengers:** {r['passenger_count']}")
                        st.write(f"**Purpose:** {r['purpose']}")
                        if r["special_request"]:
                            st.write(f"**Special Request:** {r['special_request']}")

                        if r["status"] in ("Approved", "On Trip", "Completed"):
                            approved_time = r["time_of_travel"] if is_blank(r.get("approved_time")) else r.get("approved_time")
                            time_note = (
                                f" (rescheduled from {r['time_of_travel']})"
                                if not is_blank(r.get("approved_time")) and r.get("approved_time") != r["time_of_travel"]
                                else ""
                            )
                            st.markdown(f"""
                            <div class="driver-box">
                            🚘 <b>Driver:</b> {fmt(r['driver_name'], 'TBD')} &nbsp;|&nbsp;
                            📞 <b>Contact:</b> {fmt(r['driver_contact'], 'TBD')} &nbsp;|&nbsp;
                            🔢 <b>Vehicle No.:</b> {fmt(r['vehicle_number'], 'TBD')} &nbsp;|&nbsp;
                            🕒 <b>Approved Departure Time:</b> {approved_time}{time_note}
                            </div>
                            """, unsafe_allow_html=True)
                            if not is_blank(r.get("admin_note")):
                                st.caption(f"📝 Admin Note: {r['admin_note']}")

                        if r["status"] in ("On Trip", "Completed"):
                            st.write(f"**Gate Out (Actual Exit):** {fmt(r.get('actual_exit_time'))}  |  **Start KM:** {fmt(r.get('start_km'))}")
                        if r["status"] == "Completed":
                            st.write(f"**Gate In (Actual Return):** {fmt(r.get('actual_return_time'))}  |  **End KM:** {fmt(r.get('end_km'))}  |  **Total KM:** {fmt(r.get('total_km'))}")

                        if r["status"] == "Rejected":
                            st.error("This request was rejected by the admin.")
                            if not is_blank(r.get("admin_note")):
                                st.caption(f"📝 Reason: {r['admin_note']}")

# =========================================================
# 9. SECURITY OFFICER DASHBOARD — Vehicle Gate In / Gate Out Panel
# =========================================================
elif user["role"] == "security_officer":
    company_header("🛡️ Security Officer Dashboard — Vehicle Gate Panel")
    st.caption(f"Logged in as {user['full_name']} — Security Officer")

    tab_out, tab_in = st.tabs(["🚦 Ready to Depart (Approved Trips)", "🔁 Currently On Trip (Inbound Vehicles)"])

    # ---------------- TAB 1: Ready to Depart ----------------
    with tab_out:
        st.subheader("Approved Trips Awaiting Gate Out")
        with st.spinner("Loading approved trips..."):
            ready_df = fetch_requisitions_by_status("Approved")

        if ready_df.empty:
            st.info("No trips are currently approved and waiting to depart.")
        else:
            for _, r in ready_df.iterrows():
                approved_time = r["time_of_travel"] if is_blank(r.get("approved_time")) else r.get("approved_time")
                with st.expander(f"🟢 {r['applicant_name']} ({r['department']}) → {r['destination']}  |  Vehicle: {fmt(r['vehicle_number'], 'N/A')}"):
                    c1, c2 = st.columns(2)
                    with c1:
                        st.write(f"**Applicant Name:** {r['applicant_name']}")
                        st.write(f"**Department:** {r['department']}")
                        st.write(f"**Vehicle:** {fmt(r['vehicle_number'], 'N/A')} ({r['vehicle_type']})")
                    with c2:
                        st.write(f"**Destination:** {r['destination']}")
                        st.write(f"**Requested Time:** {r['date_of_travel']} at {r['time_of_travel']}")
                        st.write(f"**Admin Approved Time:** {approved_time}")
                    if not is_blank(r.get("admin_note")):
                        st.caption(f"📝 Admin Notes: {r['admin_note']}")

                    with st.form(f"gateout_{r['request_id']}"):
                        start_km = st.number_input("Start KM (Odometer Reading) *", min_value=0.0, step=1.0,
                                                     format="%.1f", key=f"skm_{r['request_id']}")
                        depart_clicked = st.form_submit_button("🚦 Gate Out / Depart", type="primary", use_container_width=True)

                    if depart_clicked:
                        try:
                            update_requisition(r["request_id"], {
                                "start_km": float(start_km),
                                "actual_exit_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                "status": "On Trip",
                            })
                            st.success(f"✅ Gate Out recorded for {r['applicant_name']} — vehicle is now On Trip.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ Failed to record Gate Out: {e}")

    # ---------------- TAB 2: Currently On Trip ----------------
    with tab_in:
        st.subheader("Vehicles Currently Outside the Gate")
        with st.spinner("Loading active trips..."):
            ontrip_df = fetch_requisitions_by_status("On Trip")

        if ontrip_df.empty:
            st.info("No vehicles are currently on a trip.")
        else:
            for _, r in ontrip_df.iterrows():
                with st.expander(f"🔵 {r['applicant_name']} ({r['department']}) → {r['destination']}  |  Vehicle: {fmt(r['vehicle_number'], 'N/A')}"):
                    c1, c2 = st.columns(2)
                    with c1:
                        st.write(f"**Applicant Name:** {r['applicant_name']}")
                        st.write(f"**Vehicle:** {fmt(r['vehicle_number'], 'N/A')} ({r['vehicle_type']})")
                        st.write(f"**Destination:** {r['destination']}")
                    with c2:
                        st.write(f"**Gate Out Time:** {fmt(r.get('actual_exit_time'))}")
                        st.write(f"**Start KM:** {fmt(r.get('start_km'))}")

                    start_km_val = 0.0 if is_blank(r.get("start_km")) else float(r.get("start_km"))
                    with st.form(f"gatein_{r['request_id']}"):
                        end_km = st.number_input(
                            "End KM (Odometer Reading) *", min_value=start_km_val, step=1.0, format="%.1f",
                            help=f"Must be greater than or equal to Start KM ({start_km_val:.1f}).",
                            key=f"ekm_{r['request_id']}",
                        )
                        return_clicked = st.form_submit_button("🏁 Gate In / Complete", type="primary", use_container_width=True)

                    if return_clicked:
                        if end_km < start_km_val:
                            st.error("End KM cannot be less than Start KM.")
                        else:
                            total_km = round(float(end_km) - start_km_val, 1)
                            try:
                                update_requisition(r["request_id"], {
                                    "end_km": float(end_km),
                                    "total_km": total_km,
                                    "actual_return_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                    "status": "Completed",
                                })
                                st.success(f"✅ Gate In recorded for {r['applicant_name']}. Total distance: **{total_km} KM**")
                                st.rerun()
                            except Exception as e:
                                st.error(f"❌ Failed to record Gate In: {e}")

# =========================================================
# 10. ADMIN DASHBOARD
# =========================================================
elif user["role"] == "admin":
    company_header("🔐 Admin Dashboard")
    st.caption(f"Logged in as {user['full_name']} — Admin")

    tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
        "⏳ Pending User Approvals", "👥 All Users", "🚗 Pending Requisitions",
        "📊 Analytics", "📁 All Requisitions & Export", "🚘 Manage Drivers & Vehicles",
        "🕒 Duty Tracker & Analytics",
    ])

    # ---------------- TAB 1: Pending User Approvals ----------------
    with tab1:
        st.subheader("New Account Requests")
        with st.spinner("Loading users..."):
            users_df = fetch_all_users()
        pending_users = users_df[users_df["status"] == "Pending"] if not users_df.empty else users_df

        if pending_users.empty:
            st.success("🎉 No pending user registrations.")
        else:
            for _, u in pending_users.iterrows():
                role_tag = "🛡️ Security Officer" if u.get("role") == "security_officer" else "👤 Employee"
                with st.expander(f"{role_tag} — {u['full_name']} (@{u['username']})"):
                    if u.get("role") == "security_officer":
                        st.write("**Role Requested:** Security Officer")
                    else:
                        st.write(f"**Designation:** {u.get('designation', '')}")
                        st.write(f"**Employee ID:** {u.get('employee_id', '')}")
                        st.write(f"**Department:** {u.get('department', '')}")
                        st.write(f"**Mobile:** {u.get('mobile', '')}")
                    st.write(f"**Requested on:** {u.get('created_at', '')}")
                    c1, c2 = st.columns(2)
                    if c1.button("✅ Approve", key=f"appr_{u['username']}", type="primary", use_container_width=True):
                        update_user(u["username"], {"status": "Approved"})
                        st.success(f"{u['full_name']} approved.")
                        st.rerun()
                    if c2.button("❌ Reject", key=f"rej_{u['username']}", use_container_width=True):
                        update_user(u["username"], {"status": "Rejected"})
                        st.warning(f"{u['full_name']} rejected.")
                        st.rerun()

    # ---------------- TAB 2: All Users (management) ----------------
    with tab2:
        st.subheader("All User Accounts")
        if users_df.empty:
            st.info("No users yet.")
        else:
            display_cols = ["username", "full_name", "designation", "employee_id", "department",
                             "mobile", "role", "status", "created_at"]
            users_display = users_df[[c for c in display_cols if c in users_df.columns]].copy()
            if "role" in users_display.columns:
                users_display["role"] = users_display["role"].map(lambda r: ROLE_DISPLAY.get(r, r))
            st.dataframe(users_display, use_container_width=True, hide_index=True, height=300)

            st.markdown("##### Change a user's role or status")
            approved_usernames = users_df["username"].tolist()
            sel_user = st.selectbox("Select username", approved_usernames)
            if sel_user:
                rec = users_df[users_df["username"] == sel_user].iloc[0]
                c1, c2, c3 = st.columns(3)
                with c1:
                    current_role = rec.get("role", "user")
                    new_role = st.selectbox("Role", ROLE_OPTIONS,
                                             index=ROLE_OPTIONS.index(current_role) if current_role in ROLE_OPTIONS else 0,
                                             format_func=lambda r: ROLE_DISPLAY.get(r, r))
                with c2:
                    new_status = st.selectbox("Status", USER_STATUS_OPTIONS,
                                               index=USER_STATUS_OPTIONS.index(rec.get("status", "Pending")) if rec.get("status") in USER_STATUS_OPTIONS else 0)
                with c3:
                    st.write("")
                    st.write("")
                    if st.button("💾 Save Changes", use_container_width=True):
                        update_user(sel_user, {"role": new_role, "status": new_status})
                        st.success(f"Updated {sel_user}.")
                        st.rerun()

                st.markdown("---")
                st.markdown("##### 🗑️ Delete User (e.g. employee left the company)")
                if sel_user == user["username"]:
                    st.warning("You can't delete your own currently logged-in account.")
                else:
                    st.caption(
                        f"This permanently removes **{rec.get('full_name', sel_user)}** (@{sel_user}) from the system. "
                        "Their past requisition history will remain in **All Requisitions & Export**, but they will "
                        "no longer be able to log in or submit new requests. This cannot be undone."
                    )
                    confirm_delete = st.checkbox(
                        f"I understand this will permanently delete @{sel_user}'s account.",
                        key=f"confirm_del_{sel_user}",
                    )
                    if st.button("🗑️ Delete This User", type="primary", disabled=not confirm_delete,
                                 use_container_width=True, key=f"del_btn_{sel_user}"):
                        try:
                            delete_user(sel_user)
                            st.success(f"@{sel_user} has been deleted.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ Failed to delete user: {e}")

    # ---------------- TAB 3: Pending Requisitions ----------------
    with tab3:
        st.subheader("Requisitions Awaiting Action")
        with st.spinner("Loading requisitions..."):
            df_all = fetch_all_requisitions()
        pending_df = df_all[df_all["status"] == "Pending"] if not df_all.empty else df_all

        # Fetched once for this tab render and shared across every pending-request
        # card below, so each dropdown reflects the same up-to-date driver/vehicle list.
        drivers_df = fetch_all_drivers()
        vehicles_df = fetch_all_vehicles()
        driver_contact_map = dict(zip(drivers_df["driver_name"], drivers_df["driver_contact"])) if not drivers_df.empty else {}
        driver_options = ["— Select Driver —"] + drivers_df["driver_name"].tolist() if not drivers_df.empty else []
        vehicle_options = ["— Select Vehicle —"] + vehicles_df["vehicle_number"].tolist() if not vehicles_df.empty else []

        if drivers_df.empty or vehicles_df.empty:
            st.warning(
                "⚠️ No drivers and/or vehicles are registered yet. Add them under the "
                "**🚘 Manage Drivers & Vehicles** tab before you can approve requests."
            )

        if pending_df.empty:
            st.success("🎉 No pending requisitions — all caught up!")
        else:
            for _, r in pending_df.iterrows():
                with st.expander(f"🟡 {r['request_id']} — {r['applicant_name']} ({r['department']}) → {r['destination']}"):
                    c1, c2 = st.columns(2)
                    with c1:
                        st.write(f"**Mobile:** {r['mobile_number']}")
                        st.write(f"**Date/Time:** {r['date_of_travel']} at {r['time_of_travel']}")
                        st.write(f"**Passengers:** {r['passenger_count']}")
                    with c2:
                        st.write(f"**Vehicle Type:** {r['vehicle_type']}")
                        st.write(f"**Purpose:** {r['purpose']}")
                        st.write(f"**Special Request:** {r['special_request'] or '—'}")

                    # These two selects live OUTSIDE the form on purpose: widgets inside
                    # an st.form don't rerun the script until submit, so picking a driver
                    # wouldn't reveal their contact number until after clicking Approve.
                    # Outside the form, the contact updates the instant a driver is chosen.
                    d1, d2 = st.columns(2)
                    with d1:
                        selected_driver = st.selectbox(
                            "Driver Name", driver_options or ["No drivers available"],
                            key=f"drv_{r['request_id']}", disabled=not driver_options,
                        )
                    with d2:
                        bound_contact = driver_contact_map.get(selected_driver, "")
                        st.text_input("Driver Contact (auto-filled)", value=bound_contact, disabled=True,
                                      key=f"dc_disp_{r['request_id']}")
                    selected_vehicle = st.selectbox(
                        "Vehicle Number", vehicle_options or ["No vehicles available"],
                        key=f"veh_{r['request_id']}", disabled=not vehicle_options,
                    )

                    with st.form(f"action_{r['request_id']}"):
                        try:
                            default_time = datetime.strptime(r["time_of_travel"], "%H:%M").time()
                        except (ValueError, TypeError):
                            default_time = datetime.now().time()
                        approved_time = st.time_input(
                            "Approved Departure Time",
                            value=default_time,
                            help=f"Originally requested for {r['time_of_travel']}. Adjust if rescheduling.",
                            key=f"atime_{r['request_id']}",
                        )
                        admin_note = st.text_area(
                            "Admin Note / Remarks (optional)",
                            placeholder="e.g., Rescheduled due to vehicle availability, or reason for rejection",
                            key=f"note_{r['request_id']}",
                        )

                        b1, b2 = st.columns(2)
                        approve_clicked = b1.form_submit_button("✅ Approve", type="primary", use_container_width=True)
                        reject_clicked = b2.form_submit_button("❌ Reject", use_container_width=True)

                    if approve_clicked or reject_clicked:
                        new_status = "Approved" if approve_clicked else "Rejected"
                        driver_ready = driver_options and selected_driver != "— Select Driver —"
                        vehicle_ready = vehicle_options and selected_vehicle != "— Select Vehicle —"
                        if approve_clicked and not (driver_ready and vehicle_ready):
                            st.error("Please select a Driver and a Vehicle before approving.")
                        else:
                            try:
                                updates = {
                                    "status": new_status,
                                    "driver_name": selected_driver if approve_clicked else "",
                                    "driver_contact": driver_contact_map.get(selected_driver, "") if approve_clicked else "",
                                    "vehicle_number": selected_vehicle if approve_clicked else "",
                                    "approved_by": user["full_name"],
                                    "action_timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                    "admin_note": admin_note.strip(),
                                }
                                if approve_clicked:
                                    updates["approved_time"] = approved_time.strftime("%H:%M")
                                update_requisition(r["request_id"], updates)
                                st.success(f"Request {r['request_id']} marked as {new_status}.")
                                st.rerun()
                            except Exception as e:
                                st.error(f"❌ Update failed: {e}")

    # ---------------- TAB 4: Analytics ----------------
    with tab4:
        st.subheader("📊 Visual Analytics")
        if df_all.empty:
            st.info("No data yet.")
        else:
            m1, m2, m3, m4, m5, m6 = st.columns(6)
            m1.metric("Total Requests", len(df_all))
            m2.metric("🟡 Pending", int((df_all["status"] == "Pending").sum()))
            m3.metric("🟢 Approved", int((df_all["status"] == "Approved").sum()))
            m4.metric("🔵 On Trip", int((df_all["status"] == "On Trip").sum()))
            m5.metric("✅ Completed", int((df_all["status"] == "Completed").sum()))
            m6.metric("🔴 Rejected", int((df_all["status"] == "Rejected").sum()))

            c1, c2 = st.columns(2)
            with c1:
                dept_counts = df_all["department"].value_counts().reset_index()
                dept_counts.columns = ["Department", "Requests"]
                fig1 = px.bar(dept_counts, x="Department", y="Requests", color="Department", text="Requests",
                              title="Department-wise Vehicle Usage")
                fig1.update_layout(showlegend=False, height=380)
                st.plotly_chart(fig1, use_container_width=True)
            with c2:
                status_counts = df_all["status"].value_counts().reset_index()
                status_counts.columns = ["Status", "Count"]
                fig2 = px.pie(status_counts, names="Status", values="Count", hole=0.45, title="Status Breakdown",
                              color="Status",
                              color_discrete_map={"Approved": "#28a745", "Rejected": "#dc3545", "Pending": "#ffc107",
                                                   "On Trip": "#0d6efd", "Completed": "#17a673"})
                fig2.update_layout(height=380)
                st.plotly_chart(fig2, use_container_width=True)

            if "total_km" in df_all.columns and df_all["total_km"].notna().any():
                total_km_all = pd.to_numeric(df_all["total_km"], errors="coerce").dropna().sum()
                st.metric("🛣️ Total KM Covered (Completed Trips)", f"{total_km_all:.1f} KM")

            df_all["_dt"] = pd.to_datetime(df_all["date_of_travel"], errors="coerce")
            monthly = df_all.dropna(subset=["_dt"]).copy()
            if not monthly.empty:
                monthly["Month"] = monthly["_dt"].dt.to_period("M").astype(str)
                monthly_counts = monthly.groupby("Month").size().reset_index(name="Requests")
                fig3 = px.line(monthly_counts, x="Month", y="Requests", markers=True, title="Monthly Request Trend")
                fig3.update_layout(height=350)
                st.plotly_chart(fig3, use_container_width=True)

    # ---------------- TAB 5: All Requisitions + Filters + Export ----------------
    with tab5:
        st.subheader("📁 All Requisitions — Search, Filter & Export")
        if df_all.empty:
            st.info("No data yet.")
        else:
            f1, f2, f3 = st.columns(3)
            with f1:
                dept_filter = st.multiselect("Department", sorted(df_all["department"].dropna().unique().tolist()))
            with f2:
                status_filter = st.multiselect("Status", REQ_STATUS_OPTIONS)
            with f3:
                dest_filter = st.text_input("Destination contains")

            df_all["_dt"] = pd.to_datetime(df_all["date_of_travel"], errors="coerce")
            min_d = df_all["_dt"].min()
            max_d = df_all["_dt"].max()
            default_start = min_d.date() if pd.notnull(min_d) else date.today()
            default_end = max_d.date() if pd.notnull(max_d) else date.today()
            date_range = st.date_input("Date of Travel range", value=(default_start, default_end))

            filtered = df_all.copy()
            if dept_filter:
                filtered = filtered[filtered["department"].isin(dept_filter)]
            if status_filter:
                filtered = filtered[filtered["status"].isin(status_filter)]
            if dest_filter:
                filtered = filtered[filtered["destination"].str.contains(dest_filter, case=False, na=False)]
            if isinstance(date_range, tuple) and len(date_range) == 2:
                start_d, end_d = date_range
                filtered = filtered[(filtered["_dt"] >= pd.Timestamp(start_d)) & (filtered["_dt"] <= pd.Timestamp(end_d))]

            filtered_display = filtered.drop(columns=["_dt"], errors="ignore")
            st.dataframe(filtered_display, use_container_width=True, hide_index=True, height=340)

            filters_summary = (
                f"Departments: {', '.join(dept_filter) if dept_filter else 'All'} | "
                f"Status: {', '.join(status_filter) if status_filter else 'All'} | "
                f"Destination filter: {dest_filter or 'None'}"
            )

            colA, colB = st.columns(2)
            with colA:
                excel_bytes = build_excel_report(filtered_display)
                st.download_button("⬇️ Export to Excel (.xlsx)", data=excel_bytes,
                                    file_name="vehicle_requisition_report.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True)
            with colB:
                pdf_bytes = build_pdf_report(filtered_display, filters_summary)
                st.download_button("⬇️ Export to PDF (.pdf)", data=pdf_bytes,
                                    file_name="vehicle_requisition_report.pdf",
                                    mime="application/pdf",
                                    use_container_width=True)

    # ---------------- TAB 6: Manage Drivers & Vehicles ----------------
    with tab6:
        st.subheader("🚘 Manage Drivers & Vehicles")
        st.caption("These lists power the Driver and Vehicle dropdowns admins use when approving requisitions.")

        dcol, vcol = st.columns(2)

        # ---- Drivers ----
        with dcol:
            st.markdown("##### 👨‍✈️ Drivers")
            with st.form("add_driver_form", clear_on_submit=True):
                new_driver_name = st.text_input("Driver Name *")
                new_driver_contact = st.text_input("Driver Contact *", placeholder="017XXXXXXXX")
                add_driver_clicked = st.form_submit_button("➕ Add Driver", type="primary", use_container_width=True)

            if add_driver_clicked:
                if not new_driver_name.strip() or not new_driver_contact.strip():
                    st.error("Both Driver Name and Driver Contact are required.")
                else:
                    try:
                        add_driver(new_driver_name.strip(), new_driver_contact.strip())
                        st.success(f"✅ Driver '{new_driver_name.strip()}' added.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Failed to add driver: {e}")

            st.markdown("###### Current Drivers")
            drivers_df_mgmt = fetch_all_drivers()
            if drivers_df_mgmt.empty:
                st.info("No drivers added yet.")
            else:
                for _, d in drivers_df_mgmt.iterrows():
                    r1, r2 = st.columns([4, 1])
                    r1.write(f"**{d['driver_name']}** — {d['driver_contact']}")
                    if r2.button("🗑️", key=f"del_drv_{d['id']}", help="Delete this driver"):
                        try:
                            delete_driver(d["id"])
                            st.success(f"Deleted driver '{d['driver_name']}'.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ Failed to delete driver: {e}")

        # ---- Vehicles ----
        with vcol:
            st.markdown("##### 🚐 Vehicles")
            with st.form("add_vehicle_form", clear_on_submit=True):
                new_vehicle_number = st.text_input("Vehicle Number *", placeholder="e.g. DHK-METRO-GA-1234")
                add_vehicle_clicked = st.form_submit_button("➕ Add Vehicle", type="primary", use_container_width=True)

            if add_vehicle_clicked:
                if not new_vehicle_number.strip():
                    st.error("Vehicle Number is required.")
                else:
                    try:
                        add_vehicle(new_vehicle_number.strip())
                        st.success(f"✅ Vehicle '{new_vehicle_number.strip()}' added.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Failed to add vehicle: {e}")

            st.markdown("###### Current Vehicles")
            vehicles_df_mgmt = fetch_all_vehicles()
            if vehicles_df_mgmt.empty:
                st.info("No vehicles added yet.")
            else:
                for _, v in vehicles_df_mgmt.iterrows():
                    r1, r2 = st.columns([4, 1])
                    r1.write(f"**{v['vehicle_number']}**")
                    if r2.button("🗑️", key=f"del_veh_{v['id']}", help="Delete this vehicle"):
                        try:
                            delete_vehicle(v["id"])
                            st.success(f"Deleted vehicle '{v['vehicle_number']}'.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ Failed to delete vehicle: {e}")

    # ---------------- TAB 7: Duty Tracker & Analytics ----------------
    with tab7:
        st.subheader("🕒 Vehicle & Driver Duty Tracker & Analytics Dashboard")
        st.caption(
            "Filter any custom date/time window plus a specific vehicle or driver to see live duty "
            "duration, distance, and trip counts — with CSV / Excel / PDF export."
        )

        # ---- Reuse the same requisitions dataset already fetched in Tab 3/4 above ----
        # `df_all` was populated by fetch_all_requisitions() earlier in the Admin
        # Dashboard block, so we don't hit Supabase again here. Swap this for your
        # own DataFrame if you wire this tab up standalone.
        duty_df_raw = df_all.copy()

        if duty_df_raw.empty:
            st.info("No requisition data yet — the duty tracker will populate once trips are logged.")
        else:
            # -------------------------------------------------------------
            # STEP 0 — Render every filter widget FIRST, before any variable
            # derived from them is used. This avoids NameError from reading
            # a widget's value before Streamlit has actually created it.
            # -------------------------------------------------------------
            st.markdown("##### 🔎 Filters")
            fc1, fc2 = st.columns(2)
            with fc1:
                st.markdown("**Start of Range**")
                filter_start_date = st.date_input("Start Date", value=date.today(), key="duty_start_date")
                filter_start_time = st.time_input("Start Time", value=datetime.strptime("07:00", "%H:%M").time(),
                                                   key="duty_start_time")
            with fc2:
                st.markdown("**End of Range**")
                filter_end_date = st.date_input("End Date", value=date.today() + timedelta(days=1), key="duty_end_date")
                filter_end_time = st.time_input("End Time", value=datetime.strptime("06:59", "%H:%M").time(),
                                                 key="duty_end_time")

            # -------------------------------------------------------------
            # STEP 1 — Build real start/end datetimes for every trip, as
            # UTC-aware timestamps. Supabase/PostgREST returns timestamptz
            # columns as ISO 8601 strings — sometimes with an explicit
            # offset, sometimes without — so `utc=True` normalizes BOTH
            # cases onto a single tz-aware ("datetime64[us, UTC]" / similar)
            # dtype. Without this, mixed naive/aware values make pandas
            # infer a naive dtype for one column and an aware dtype for
            # another, and mixing the two later raises exactly the
            # "Invalid comparison between dtype=datetime64[...,UTC] and
            # datetime" TypeError.
            # -------------------------------------------------------------
            # `format="mixed"` matters as much as `utc=True` here: Supabase/
            # PostgREST timestamptz values can come back with or without
            # fractional seconds, or with a trailing "Z" vs "+00:00" offset,
            # row to row. Without `format="mixed"`, pandas infers a single
            # format from the first non-null value and silently coerces every
            # differently-shaped row to NaT (even with errors="coerce") —
            # which then makes real trips vanish from the tracker with no
            # error at all. "mixed" parses each value independently.
            duty_df_raw["_start_dt"] = pd.to_datetime(
                duty_df_raw.get("actual_exit_time"), errors="coerce", utc=True, format="mixed"
            )
            duty_df_raw["_end_dt"] = pd.to_datetime(
                duty_df_raw.get("actual_return_time"), errors="coerce", utc=True, format="mixed"
            )

            # A trip only has meaningful "duty duration" once Security has
            # logged a Gate Out (actual_exit_time). If Gate In hasn't
            # happened yet (still "On Trip"), treat "now" (in UTC, to match
            # the column's dtype) as the running end time so in-progress
            # duty shows up too. Assigning a naive Timestamp into a UTC-aware
            # column is exactly what triggers the
            # "Invalid value ... for dtype 'datetime64[us, UTC]'" TypeError,
            # so we must assign an equally tz-aware Timestamp here.
            still_out_mask = duty_df_raw["_start_dt"].notna() & duty_df_raw["_end_dt"].isna()
            duty_df_raw.loc[still_out_mask, "_end_dt"] = pd.Timestamp.now(tz="UTC")

            # Only rows that actually left the gate are real "duty" records.
            duty_base = duty_df_raw[duty_df_raw["_start_dt"].notna()].copy()

            # -------------------------------------------------------------
            # STEP 2 — Combine the date/time widgets into naive datetimes,
            # then localize them to UTC so they can be compared directly
            # against the tz-aware `_start_dt` / `_end_dt` columns above.
            # (If your admin users think in a local timezone rather than
            # UTC, swap "UTC" below for that zone, e.g. "Asia/Dhaka", and
            # pandas will convert correctly at comparison time.)
            # -------------------------------------------------------------
            range_start = pd.Timestamp(datetime.combine(filter_start_date, filter_start_time)).tz_localize("UTC")
            range_end = pd.Timestamp(datetime.combine(filter_end_date, filter_end_time)).tz_localize("UTC")

            if range_start >= range_end:
                st.error("⚠️ The start date/time must be earlier than the end date/time.")
                st.stop()

            fc3, fc4 = st.columns(2)
            with fc3:
                vehicle_choices = ["All Vehicles"] + sorted(
                    v for v in duty_base["vehicle_number"].dropna().unique().tolist() if v
                )
                duty_vehicle_filter = st.selectbox("Vehicle Selection", vehicle_choices, key="duty_vehicle_filter")
            with fc4:
                driver_choices = ["All Drivers"] + sorted(
                    d for d in duty_base["driver_name"].dropna().unique().tolist() if d
                )
                duty_driver_filter = st.selectbox("Driver Selection", driver_choices, key="duty_driver_filter")

            # -------------------------------------------------------------
            # STEP 3 — Apply the date/time window + vehicle/driver filters.
            # A trip is included if its duty window OVERLAPS the selected
            # range at all (not just if it starts inside it) — this correctly
            # captures overnight duties like "07:00 today to 06:59 tomorrow".
            # Both sides are now UTC-aware, so this comparison is safe.
            # -------------------------------------------------------------
            duty_filtered = duty_base[
                (duty_base["_start_dt"] <= range_end) & (duty_base["_end_dt"] >= range_start)
            ].copy()

            if duty_vehicle_filter != "All Vehicles":
                duty_filtered = duty_filtered[duty_filtered["vehicle_number"] == duty_vehicle_filter]
            if duty_driver_filter != "All Drivers":
                duty_filtered = duty_filtered[duty_filtered["driver_name"] == duty_driver_filter]

            # -------------------------------------------------------------
            # STEP 4 — Derived fields: duty duration in hours, total KM.
            # -------------------------------------------------------------
            duty_filtered["_duration_hrs"] = (
                (duty_filtered["_end_dt"] - duty_filtered["_start_dt"]).dt.total_seconds() / 3600.0
            ).round(2)
            duty_filtered["_km"] = pd.to_numeric(duty_filtered.get("total_km"), errors="coerce").fillna(0.0)

            st.markdown("---")
            st.markdown("##### 📌 Summary")

            total_km = float(duty_filtered["_km"].sum())
            total_duration_hrs = float(duty_filtered["_duration_hrs"].sum())
            total_trips = int(len(duty_filtered))
            duration_h = int(total_duration_hrs)
            duration_m = int(round((total_duration_hrs - duration_h) * 60))

            # "Active Driver & Assigned Vehicle" — meaningful only when the
            # filters have narrowed things down to one driver/vehicle; with
            # "All Drivers"/"All Vehicles" selected we instead surface the
            # busiest one within the filtered window.
            if duty_driver_filter != "All Drivers":
                active_driver_label = duty_driver_filter
            elif not duty_filtered.empty and duty_filtered["driver_name"].notna().any():
                active_driver_label = duty_filtered["driver_name"].value_counts().idxmax()
            else:
                active_driver_label = "—"

            if duty_vehicle_filter != "All Vehicles":
                active_vehicle_label = duty_vehicle_filter
            elif not duty_filtered.empty and duty_filtered["vehicle_number"].notna().any():
                active_vehicle_label = duty_filtered["vehicle_number"].value_counts().idxmax()
            else:
                active_vehicle_label = "—"

            k1, k2, k3, k4 = st.columns(4)
            k1.metric("🛣️ Total Distance", f"{total_km:.1f} KM")
            k2.metric("⏱️ Total Duty Duration", f"{duration_h}h {duration_m}m")
            k3.metric("🚗 Total Trips", total_trips)
            k4.metric("👨‍✈️ Active Driver / Vehicle", f"{active_driver_label} / {active_vehicle_label}")

            # -------------------------------------------------------------
            # STEP 5 — Detailed table:
            # [Vehicle No, Driver Name, Start Time, End Time, Total KM,
            #  Duty Duration (Hours), Route / Purpose]
            # -------------------------------------------------------------
            st.markdown("---")
            st.markdown("##### 📋 Detailed Duty Log")

            if duty_filtered.empty:
                st.info("No trips match the selected filters.")
                detail_display = pd.DataFrame(columns=DUTY_TRACKER_DISPLAY_COLS)
            else:
                detail_display = pd.DataFrame({
                    "Vehicle No": duty_filtered["vehicle_number"].map(lambda v: fmt(v, "—")),
                    "Driver Name": duty_filtered["driver_name"].map(lambda v: fmt(v, "—")),
                    "Start Time": duty_filtered["_start_dt"].dt.strftime("%Y-%m-%d %H:%M"),
                    "End Time": duty_filtered["_end_dt"].dt.strftime("%Y-%m-%d %H:%M"),
                    "Total KM": duty_filtered["_km"].round(1),
                    "Duty Duration (Hrs)": duty_filtered["_duration_hrs"],
                    "Route / Purpose": duty_filtered["destination"].fillna("").astype(str)
                                        + " — " + duty_filtered["purpose"].fillna("").astype(str),
                }).reset_index(drop=True)
                st.dataframe(detail_display, use_container_width=True, hide_index=True, height=340)

            # -------------------------------------------------------------
            # STEP 6 — Multi-format export: CSV / Excel / PDF.
            # summary_metrics feeds both the Excel "Summary" sheet and the
            # PDF's KPI block — plug in any additional metrics here as needed.
            # -------------------------------------------------------------
            summary_metrics = {
                "Date/Time Range": f"{range_start.strftime('%Y-%m-%d %H:%M')} to {range_end.strftime('%Y-%m-%d %H:%M')}",
                "Vehicle Filter": duty_vehicle_filter,
                "Driver Filter": duty_driver_filter,
                "Total Distance (KM)": f"{total_km:.1f}",
                "Total Duty Duration": f"{duration_h}h {duration_m}m",
                "Total Trips": total_trips,
                "Active Driver": active_driver_label,
                "Assigned Vehicle": active_vehicle_label,
            }
            filters_summary_text = (
                f"Range: {summary_metrics['Date/Time Range']} | Vehicle: {duty_vehicle_filter} | "
                f"Driver: {duty_driver_filter}"
            )

            st.markdown("---")
            st.markdown("##### ⬇️ Export Duty Report")
            e1, e2, e3 = st.columns(3)
            with e1:
                csv_bytes = detail_display.to_csv(index=False).encode("utf-8")
                st.download_button(
                    "⬇️ Download CSV", data=csv_bytes, file_name="duty_tracker_report.csv",
                    mime="text/csv", use_container_width=True,
                )
            with e2:
                duty_excel_bytes = build_duty_tracker_excel(detail_display, summary_metrics)
                st.download_button(
                    "⬇️ Download Excel (.xlsx)", data=duty_excel_bytes, file_name="duty_tracker_report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with e3:
                duty_pdf_bytes = build_duty_tracker_pdf(detail_display, summary_metrics, filters_summary_text)
                st.download_button(
                    "⬇️ Download PDF (.pdf)", data=duty_pdf_bytes, file_name="duty_tracker_report.pdf",
                    mime="application/pdf", use_container_width=True,
                )

# =========================================================
# 11. FALLBACK — unrecognized role
# =========================================================
else:
    st.error("⚠️ Your account role is not recognized. Please contact the Admin.")
